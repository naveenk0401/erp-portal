from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
from app.services.base_service import BaseService
from app.models.finance import Expense, Invoice, PayrollRecord
from app.models.salary import Salary
from beanie import Document
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class FinanceService:
    async def get_summary(self, period: str = "year") -> Dict[str, Any]:
        """Calculate real-time financial metrics with period filtering"""
        try:
            # Filter Logic (Mocked for now, but structure enabled)
            cutoff = datetime.utcnow() - timedelta(days=365)
            if period == "month":
                cutoff = datetime.utcnow() - timedelta(days=30)
            elif period == "week":
                cutoff = datetime.utcnow() - timedelta(days=7)

            # Aggregate Invoices (Revenue) - using grand_total
            invoices = await Invoice.find(Invoice.status == "paid").to_list() # Add date filter in real query
            total_revenue = sum(inv.grand_total for inv in invoices)
            
            # Aggregate Expenses
            expenses = await Expense.find(Expense.status == "approved").to_list()
            total_expenses = sum(exp.amount for exp in expenses)
            
            # Aggregate Salaries
            salaries = await Salary.find(Salary.status == "paid").to_list()
            total_salaries = sum(sal.amount + sal.bonus - sal.deductions for sal in salaries)
            
            # Net Worth calculation
            net_worth_val = total_revenue - (total_expenses + total_salaries)
            
            return {
                "net_worth": f"₹{(net_worth_val / 10000000):.1f} Cr" if net_worth_val > 10000000 else f"₹{(net_worth_val / 100000):.1f} L",
                "net_worth_trend": 12.4, # Mock trend
                "cashflow": f"₹{(total_revenue / 10000000):.1f} Cr" if total_revenue > 10000000 else f"₹{(total_revenue / 100000):.1f} L",
                "cashflow_trend": 4.2,
                "operating_margin": f"{( (total_revenue - total_expenses) / total_revenue * 100 if total_revenue > 0 else 0):.1f}%",
                "operating_margin_trend": -1.2,
                "efficiency_indicator": "High" if total_revenue > total_expenses * 2 else "Stable",
            }
        except Exception as e:
            print(f"[FINANCE SERVICE] Error calculating summary: {str(e)}")
            return {"net_worth": "₹0.0 Cr", "net_worth_trend": 0, "cashflow": "₹0.0 Cr", "cashflow_trend": 0, "operating_margin": "0%", "operating_margin_trend": 0, "efficiency_indicator": "N/A"}

    async def get_expenses_by_category(self) -> List[Dict[str, Any]]:
        expenses = await Expense.find_all().to_list()
        cat_map = {}
        for exp in expenses:
            cat_map[exp.category] = cat_map.get(exp.category, 0) + exp.amount
        
        return [{"category": k, "amount": v} for k, v in cat_map.items()]

    async def seed_demo_data(self):
        """Seed detailed B2B invoice data"""
        if await Invoice.count() == 0:
            demo_invoices = [
                {
                    "to_company": {"name": "Stark Industries", "address": "10880 Malibu Point, CA", "gst": "27AAACS1234A1Z5"},
                    "items": [{"description": "Arc Reactor Maintenance", "quantity": 1, "rate": 12000000}],
                    "status": "paid",
                    "due_date": datetime.utcnow()
                },
                {
                    "to_company": {"name": "Wayne Enterprises", "address": "1007 Mountain Drive, Gotham", "gst": "27AAACW5678B1Z2"},
                    "items": [{"description": "Security System Upgrade", "quantity": 1, "rate": 8500000}],
                    "status": "pending",
                    "due_date": datetime.utcnow()
                },
                {
                    "to_company": {"name": "Cyberdyne Systems", "address": "18144 El Camino Real, Sunnyvale", "gst": "27AAACC9012C1Z9"},
                    "items": [{"description": "Neural Net Processor Analysis", "quantity": 2, "rate": 1750000}],
                    "status": "paid",
                    "due_date": datetime.utcnow()
                }
            ]

            for inv_data in demo_invoices:
                # Calculate totals
                subtotal = sum(i["quantity"] * i["rate"] for i in inv_data["items"])
                tax_total = subtotal * 0.18
                grand_total = subtotal + tax_total

                inv = Invoice(
                    to_company=inv_data["to_company"],
                    from_company={"name": "ERP Systems Ltd", "address": "Tech Park, Bangalore", "gst": "29ABCDE1234F1Z5"},
                    items=inv_data["items"],
                    subtotal=subtotal,
                    tax_total=tax_total,
                    grand_total=grand_total,
                    status=inv_data["status"],
                    due_date=inv_data["due_date"],
                    payment_details={"bank_name": "HDFC Bank", "account_no": "50200012345678", "ifsc": "HDFC0001234", "mode": "neft"}
                )
                await inv.insert()
            print("[FINANCE SEED] Detailed Invoices seeded")

        if await Expense.count() == 0:
            # Seed expenses (Simplified)
            expenses = [
                Expense(user=None, amount=120000, category="Cloud Infrastructure", description="AWS Monthly Bill", status="approved"),
                Expense(user=None, amount=45000, category="Marketing", description="Google Ads - Q1 Campaign", status="approved"),
            ]
            from app.models.user import User
            admin = await User.find_one({"role": "super_admin"})
            for exp in expenses:
                if admin: exp.user = admin
                await exp.insert()
            print("[FINANCE SEED] Expenses seeded")

    async def generate_pdf_bytes(self, invoice_id: str) -> io.BytesIO:
        """Generate PDF for a given invoice using ReportLab"""
        inv = await Invoice.get(invoice_id)
        if not inv:
            return None

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Custom Styles
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=24, spaceAfter=20)
        h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontSize=14, spaceAfter=10, textColor=colors.HexColor('#1e293b'))
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=10, leading=14)

        # Header (Company Info)
        elements.append(Paragraph(f"INVOICE: {inv.invoice_number}", title_style))
        elements.append(Spacer(1, 10))
        
        # Addresses
        data = [
            [Paragraph("<b>From:</b>", normal_style), Paragraph("<b>Bill To:</b>", normal_style)],
            [
                Paragraph(f"{inv.from_company.get('name')}<br/>{inv.from_company.get('address')}<br/>GSTIN: {inv.from_company.get('gst')}", normal_style),
                Paragraph(f"{inv.to_company.get('name')}<br/>{inv.to_company.get('address')}<br/>GSTIN: {inv.to_company.get('gst')}", normal_style)
            ]
        ]
        t = Table(data, colWidths=[3.5*inch, 3.5*inch])
        t.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

        # Meta Data (Date, Due Date)
        elements.append(Paragraph(f"<b>Invoice Date:</b> {inv.created_at.strftime('%d-%b-%Y')} &nbsp;&nbsp;&nbsp; <b>Due Date:</b> {inv.due_date.strftime('%d-%b-%Y')}", normal_style))
        elements.append(Spacer(1, 20))

        # Line Items Table
        item_data = [['Description', 'HSN/SAC', 'Qty', 'Rate', 'Amount']]
        for item in inv.items:
            amount = item.get('quantity', 0) * item.get('rate', 0)
            item_data.append([
                Paragraph(item.get('description', ''), normal_style),
                item.get('hsn_sac', ''),
                str(item.get('quantity', 0)),
                f"{item.get('rate', 0):,.2f}",
                f"{amount:,.2f}"
            ])
        
        # Totals
        item_data.append(['', '', '', 'Subtotal:', f"{inv.subtotal:,.2f}"])
        item_data.append(['', '', '', 'Tax (18%):', f"{inv.tax_total:,.2f}"])
        item_data.append(['', '', '', 'Grand Total:', f"{inv.grand_total:,.2f}"])

        t_items = Table(item_data, colWidths=[3*inch, 1*inch, 0.75*inch, 1.25*inch, 1.5*inch])
        t_items.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#0f172a')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ALIGN', (2,1), (-1,-1), 'RIGHT'), # Qty, Rate, Amount align right
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-4), 0.5, colors.grey), # Grid for items
            ('LINEBELOW', (0,-3), (-1,-1), 0.5, colors.grey), # Lines for totals
            ('FONTNAME', (-2,-1), (-1,-1), 'Helvetica-Bold'), # Grand total bold
        ]))
        elements.append(t_items)
        elements.append(Spacer(1, 30))

        # Payment Info & Terms
        elements.append(Paragraph("Payment Layout", h2_style))
        pay_info = inv.payment_details
        elements.append(Paragraph(f"Bank: {pay_info.get('bank_name')}<br/>Account: {pay_info.get('account_no')}<br/>IFSC: {pay_info.get('ifsc')}", normal_style))
        
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("Terms & Conditions", h2_style))
        elements.append(Paragraph(inv.terms_conditions, normal_style))
        
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("Authorized Signatory", normal_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    async def generate_audit_excel_bytes(self) -> io.BytesIO:
        """Generate Detailed Excel Audit Report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Audit"

        # Headers
        headers = ["Date", "Invoice #", "Client", "Amount", "Tax", "Total", "Status", "Due Date"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data
        invoices = await Invoice.find_all().sort("-created_at").to_list()
        row_num = 2
        for inv in invoices:
            ws.cell(row=row_num, column=1, value=inv.created_at.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=2, value=inv.invoice_number)
            ws.cell(row=row_num, column=3, value=inv.to_company.get('name', inv.dict().get('client_name', 'N/A'))) # Fallback
            ws.cell(row=row_num, column=4, value=inv.subtotal)
            ws.cell(row=row_num, column=5, value=inv.tax_total)
            ws.cell(row=row_num, column=6, value=inv.grand_total if inv.grand_total > 0 else inv.amount) # Fallback
            
            # Status styling
            status_cell = ws.cell(row=row_num, column=7, value=inv.status.upper())
            if inv.status == 'paid':
                status_cell.fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid") # Green
                status_cell.font = Font(color="166534", bold=True)
            elif inv.status == 'overdue':
                status_cell.fill = PatternFill(start_color="ffe4e6", end_color="ffe4e6", fill_type="solid") # Red
                status_cell.font = Font(color="9f1239", bold=True)
            
            ws.cell(row=row_num, column=8, value=inv.due_date.strftime('%Y-%m-%d'))
            row_num += 1

        # Adjust widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

finance_service = FinanceService()
